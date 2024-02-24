from typing import List
from openpyxl import load_workbook
from xls2xlsx import XLS2XLSX
import requests
import sqlite3
import datetime
import os
import asyncio


user_agent: dict = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36'}
url1: str = "https://www.vyatsu.ru/studentu-1/spravochnaya-informatsiya/teacher.html"


# актуализация базы данных
async def clear_db():
    pass


# получение данных с сайта
async def get_content(url: str):
    return requests.get(url, headers=user_agent)


# получение актуальных ссылок с сайта
async def get_urls(response):
    list_urls = []
    text = response.text
    
    # ищем первую ссылку
    index = text.find('href="/reports/')
    text = text[index+6:]
    
    # пока есть ссылки
    while index > 0:
        i = 0
        # пока не конец ссылки
        url = "https://www.vyatsu.ru"
        while text[i]!="\"":
            # формируем url
            url += text[i]
            i += 1
        if url[-3:] == "xls":
            temp_url = url[:-4]
            year_end = int(temp_url[-4:])
            month_end = int(temp_url[-6:-4])
            day_end = int(temp_url[-8:-6])
            date = datetime.date(year_end, month_end, day_end)
            current_date = datetime.date.today()
            if current_date < date and (date-current_date).days <= 28:
                list_urls.append(url)
                
        text = text[i:]
        # ищем новую ссылку
        index = text.find('href="/reports/')
        text = text[index+6:]   
    return list_urls


async def update_db(list_urls: List[str]):
    base = sqlite3.connect("database.db")
    cur = base.cursor()
    base.execute("CREATE TABLE IF NOT EXISTS schedule(date text, time_lesson text, name_teacher text, cabinet_number text, name_of_group_and_discipline text)")
    base.commit()
    
    # для каждой ссылки
    for url in list_urls:
        # скачивание файла
        response = requests.get(url, headers=user_agent)
        with open(url[-27:], 'wb') as file:
            file.write(response.content)
            
        # конвертируем в xlsx
        x2x = XLS2XLSX(url[-27:])
        path = url[-27:] + "x"
        x2x.to_xlsx(path)
        
        #открываем файл
        wookbook = load_workbook(path)
        worksheet = wookbook.active

        #парсим файл
        print(worksheet.max_column)
        for i in range(3, worksheet.max_column):
            name_teacher = worksheet.cell(row=2, column=i).value
            

        # удаляем оба файла
        os.remove(url[-27:])
        os.remove(path)
    base.close()
    

async def start_parsing_xlsx():
    while True:
        
        print(f"Время начала обновления: {datetime.datetime.now()}")
        
        # удаление неактуальных данных
        await clear_db()
        
        # получение данных с сайта
        response = await get_content(url1)
        
        # получение ссылок с сайта
        list_urls = await get_urls(response)
        
        # заполнение базы данных id файлов
        await update_db(list_urls)
        
        print(f"Время конца обновления: {datetime.datetime.now()}")
        
        # обновляем бд каждые 4 часа
        await asyncio.sleep(12000)


def main():
    asyncio.run(start_parsing_xlsx())
    

if __name__ == '__main__':
    asyncio.run(start_parsing_xlsx())