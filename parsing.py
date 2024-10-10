from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook
import os
import requests
import datetime
import asyncio
import logging
import traceback

from dbrequests import delete_outdated_schedules, update_schedule

logging.basicConfig(
    filename='log.txt',
    level=logging.ERROR, 
    format='%(asctime)s - %(levelname)s - %(message)s'
)

time_from_pair = {
    "1 пара": "8:20-9:50",
    "2 пара": "10:00-11:30",
    "3 пара": "11:45-13:15",
    "4 пара": "14:00-15:30",
    "5 пара": "15:45-17:15",
    "6 пара": "17:20-18:50",
    "7 пара": "18:55-20:15",
    }

user_agent = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36'}
url_site = "https://www.vyatsu.ru/studentu-1/spravochnaya-informatsiya/zanyatost-auditoriy.html"


# получение данных с сайта
async def get_content(url):
    return requests.get(url, headers=user_agent)


# получение актуальных ссылок с сайта
async def get_urls(response):
    list_urls = []
    text: str = response.text
    
    while True:
        # Ищем ссылку на отчет
        index = text.find('href="/reports/')
        if index == -1:
            break  # Выходим из цикла, если ссылки больше нет
        
        # Извлекаем URL
        start_index = index + len('href="')
        end_index = text.find('"', start_index)
        url = "https://www.vyatsu.ru" + text[start_index:end_index]
        
        # Проверяем, если ссылка на .xls
        if url.endswith(".xls"):
            try:
                # Извлекаем дату из URL
                date_str = url[-12:-4]
                last_date = datetime.date(
                    year=int(date_str[-4:]), 
                    month=int(date_str[-6:-4]), 
                    day=int(date_str[-8:-6])
                )
                
                # Проверяем дату на актуальность
                current_date = datetime.date.today()
                if current_date < last_date and (last_date - current_date).days < 60:
                    list_urls.append(url)
            except Exception as e:
                logging.error(f"{e}")
                logging.error(traceback.format_exc())

        # Продолжаем поиск следующей ссылки
        text = text[end_index:]
    return list_urls


# скачивание файла с расширением .xls
async def download(url: str) -> None:
    response = requests.get(url, headers=user_agent)
    with open(url[-25:], 'wb') as file:
        file.write(response.content)


# конвертирование файлв из .xls в .xlsx
async def convert_xls_to_xlsx(path: str) -> None:
    x2x = XLS2XLSX(path)
    x2x.to_xlsx(path + 'x')


# парсинг файла и заполнение бд данными
async def parsing_url(url: str) -> None:
    # скичивание файла
    await download(url)
            
    # конвертирование файла из .xls в .xlsx
    path = url[-25:]
    await convert_xls_to_xlsx(path)
    
    # удаляем старый файл
    os.remove(path)
    
    #открываем файл уже с расширением .xlsx
    path += 'x'
    wookbook = load_workbook(path)
    worksheet = wookbook.active

    # парсим файл
    for i in range(3, worksheet.max_column):
        # номер кабинета
        cabinet_number: str = worksheet.cell(row=2, column=i).value
        
        # по каждому столбцу
        for j in range(3, worksheet.max_row):
            # номер пары
            time_lesson: str = worksheet.cell(row=j, column=2).value.strip()
            # время пары
            time_lesson: str = time_from_pair[time_lesson]
            # дата
            date = worksheet.cell(row=j, column=1).value
            # если дата пустая, то ищем дату, зачем оставляем только число xx-xx-xx
            index_for_day = j - 1
            while date == None:
                date = worksheet.cell(row=index_for_day, column=1).value
                index_for_day -= 1
            date = date.strip()[-8:]
            
            # данные основной ячейки
            value: str = worksheet.cell(row=j, column=i).value
            
            # если что-то есть, но не нужное - удаляем
            if value != None:
                if "Резервирование" in value:
                    value = None
            
            # парсим, если что-то есть
            if value != None:
                # если несколько групп в одном кабинете
                if "\n" in value:
                    list_value: list[str] = value.split("\n")
                else:
                    list_value: list[str] = [value]
                    
                # убираем концевые пробелы
                list_value: list[str] = [value.strip() for value in list_value]
                
                # убираем вначале приписку из двух-трёх символом, если она есть
                for ii in range(len(list_value)):
                    if len(list_value[ii].split()[0]) <= 3:
                        list_value[ii] = " ".join(list_value[ii].split()[1:])

                # списки(если на паре несколько групп)
                name_of_group = []
                name_teacher = []
                name_of_discipline = []     
                
                # разделяем данные для каждой записи в ячейке(в одном кабинете может быть несколько групп)
                for value in list_value:
                    # делим строку
                    text_split = value.split()
                    
                    # если две точки есть в ласт слове, то инициалы = препод
                    if text_split[-1].count(".") == 2:
                        name_teacher.append(" ".join(text_split[-2:]))
                        text_split = text_split[:-2]
                    # иначе препода нет
                    else:
                        name_teacher.append(None)
                    
                    # если в конце названия группы запятая, то дальше идёт номер подгруппы
                    if text_split[0][-1] == ",":
                        name_of_group.append(text_split[0][:-1])
                        text_split = text_split[3:]
                    else:
                        name_of_group.append(text_split[0])
                        text_split = text_split[1:]
                        
                    # название дисциплины
                    name_of_discipline.append(" ".join(text_split))    
                
                # если на паре одна группа
                if len(list_value)==1:
                    await update_schedule(date, time_lesson, cabinet_number, name_of_group, name_teacher, name_of_discipline)
                    
                # если на паре несколько групп
                else:
                    await update_schedule(date, time_lesson, cabinet_number, name_of_group, name_teacher, name_of_discipline, many = True)
                    
            # если пары нет на это дату, время и кабинет
            else:
                await update_schedule(date, time_lesson, cabinet_number, None, None, None, empty = True)

    # удаляем файд
    os.remove(path)
        

# цикл обработки файлов
async def start_parsing_xlsx():
    while True:
        await delete_outdated_schedules()
        try:
            response = await get_content(url_site)
            urls = await get_urls(response)
            for url in urls:
                try:
                    await parsing_url(url)
                except Exception as e:
                    logging.error(f"{e}")
                    logging.error(traceback.format_exc())
        except Exception as e:
            logging.error(f"{e}")
            logging.error(traceback.format_exc())
        await asyncio.sleep(9000)


def start_parsing():
    asyncio.run(start_parsing_xlsx())


if __name__ == "__main__":
    start_parsing()